import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

workbook = Workbook()

workbook["Sheet"].title="Report Of Empolyees"

Active = workbook.active
Current = workbook['Report Of Empolyees']

data = [["Employee No","Name","Post","Salary","Status","Gender","Date Of Joining","Age","Marital Status","Qulification"],[1,"Yash","CEO",150000,"Active","Male","21 April 2024",19,"Single","HSC Passed"],[2,"Pavn","Co-Founder",100000,"Active","Male","22 April 2024",20,"Single","HSC Passed"]]
rowNo = len(data)
collumnNO = 5

def CreateDatabase():
    for i in data:
        Current.append(i)

    Current.cell(row=2,column=collumnNO).fill=PatternFill("solid",fgColor="008000")
    Current.cell(row=3,column=collumnNO).fill=PatternFill("solid",fgColor="008000")

cur_No = 2
EmployeeNo = cur_No + 1    

def Add_Employee(EmployeeNo):
    Lis = []
    
    Lis.append(EmployeeNo)

    Name = input("Enter The Name Of Employee :")
    Lis.append(Name)

    Post = input("Enter The Post Of Employee :")
    Lis.append(Post)

    Salary = int(input("Enter The Salary Of Employee :"))
    Lis.append(Salary)

    Status = "Active"
    Lis.append(Status)

    Gender = input("Enter Gender of Employee :")
    Lis.append(Gender)

    Date = input("Enter The Date of Joining :")
    Lis.append(Date)

    Age = int(input("Enter The Age Of Employee :"))
    Lis.append(Age)

    Marital_Status = input("Enter Marital Status :")
    Lis.append(Marital_Status)

    Qulatification = input("Enter Qulification Of Employee :")
    Lis.append(Qulatification)

    data.append(Lis)
    Current.append(Lis)
    Current.cell(row=EmployeeNo+1,column=collumnNO).fill=PatternFill("solid",fgColor="008000")
    
    

    # file_name = input("Enter The Name for Updated File :")
    # workbook.save(f"{file_name}.xlsx")



def Remove_Employee():
    Emp_Number = int(input("Enter The No Of Employee You Want To Remove :"))
    # data[Emp_Number][4] = "Inactive"
    data[Emp_Number][4] = "InActive"
    Current.cell(row=Emp_Number+1,column=collumnNO,value='InActive')
    Current.cell(row=Emp_Number+1,column=collumnNO).fill=PatternFill("solid",fgColor="ff0000")
    # file_name = input("Enter The Name for Updated File :")
    # workbook.save(f"{file_name}.xlsx")



def DeleteDataBase():
    import os
    file_path = "EmployeeDataBase.xlsx"
    try:
        os.remove(file_path)
        print("File deleted successfully!")
    except FileNotFoundError:
        print("Error: File not found.")
    except PermissionError:
         print("Error: You don't have permission to delete the file.")
    except OSError as e:
        print(f"Error deleting file: {e}")


def PromoteEmployee():
    Emp_Number = int(input("Enter The No Of Employee You Want To Promote :"))
    Post = input(f"Enter The Promoted Post for Employee No {Emp_Number} :")
    Salaray = int(input(f"Enter The Salary For the Employee{Emp_Number} :"))
    data[Emp_Number][2] = Post
    data[Emp_Number][3] = Salaray
    Current.cell(row=Emp_Number+1,column=3,value=Post)
    Current.cell(row=Emp_Number+1,column=4,value=Salaray)
    print(f"Employee No {Emp_Number} is Promoted To Post {Post}")
    # file_name = input("Enter The Name for Updated File :")
    # workbook.save(f"{file_name}.xlsx")

def DisplayEmployeeData():
    Emp_No = int(input('Enter the Number of employee :'))
    temp = data[Emp_No]
    temp2 = data[0]
    print(temp2[0],end=" : ")
    print(temp[0],end=" ")
    print()
    print(temp2[1],end=" : ")
    print(temp[1],end=" ")
    print()
    print(temp2[2],end=" : ")
    print(temp[2],end=" ")
    print()
    print(temp2[3],end=" : ")
    print(temp[3],end=" ")
    print()
    print(temp2[4],end=" : ")
    print(temp[4],end=" ")
    print()
    print(temp2[5],end=" : ")
    print(temp[5],end=" ")
    print()
    print(temp2[6],end=" : ")
    print(temp[6],end=" ")
    print()
    print(temp2[7],end=" : ")
    print(temp[7],end=" ")
    print()
    print(temp2[8],end=" : ")
    print(temp[8],end=" ")
    print()
    print(temp2[8],end=" : ")
    print(temp[8],end=" ")
    print()

def DisplayWholeDataSet():
    for i in range(1,len(data)):
        temp = data[i]
        temp2 = data[0]
        print(temp2[0],end=" : ")
        print(temp[0],end=" ")
        print()
        print(temp2[1],end=" : ")
        print(temp[1],end=" ")
        print()
        print(temp2[2],end=" : ")
        print(temp[2],end=" ")
        print()
        print(temp2[3],end=" : ")
        print(temp[3],end=" ")
        print()
        print(temp2[4],end=" : ")
        print(temp[4],end=" ")
        print()
        print(temp2[5],end=" : ")
        print(temp[5],end=" ")
        print()
        print(temp2[6],end=" : ")
        print(temp[6],end=" ")
        print()
        print(temp2[7],end=" : ")
        print(temp[7],end=" ")
        print()
        print(temp2[8],end=" : ")
        print(temp[8],end=" ")
        print()
        print(temp2[8],end=" : ")
        print(temp[8],end=" ")
        print()
        print('---------------------------------------------------------------------------')


import time
timestamp =time.strftime('%H:%M:%S')
#print(timestamp)
timestamp = time.strftime('%H')
# print(timestamp)
timestamp = time.strftime('%M')
# print(timestamp)
timestamp = time.strftime('%S')
# print(timestamp)

checker = int(time.strftime('%H'))
if(checker<12):
    print("😊😊😊😊😊😊😊😊😊😊😊Good Morning😊😊😊😊😊😊😊😊😊😊😊😊")
    

elif(12<= checker <18):
     print("😊😊😊😊😊😊😊😊😊😊😊😊Good Afternoon😊😊😊😊😊😊😊😊😊😊😊😊")

elif(18< checker <21):
     print("😊😊😊😊😊😊😊😊😊😊😊😊Good Evening😊😊😊😊😊😊😊😊😊😊😊😊")

print()
print()
print(">>>>>>>>>>>>>>>Welcome To Our Project<<<<<<<<<<<<<<<<<<<<<<")
print()
print("Made By :\n     109207 Pavan Gavit\n     108209 Yash Maske\n     108216 Chandrkant Thakre\n     108221 Shubhankar Jakate\n")
print()
print("This Program Is Able To Create A Data Base Of Employee and Can Perform Some Operations")
print()
print()
print("^^^^^^^^^^^^^^^^^^Hit Enter To Continue^^^^^^^^^^^^^^^^^^^")
EnterHitter = input()

print("Please Select The Operation No You Want To Perform On DataBase ")
print()
print()
print()
print("Options Are\n1.Create A Database\n2.Add An Employee\n3.Promote An Employee\n4.Remove An Employee\n5.Display Information\n6.Display Whole Data Set\n7.Delete DataBase\n8.Exit\n")
print()
print('These Are The Operations Available')

# choice = int(input("Enter The No Of Operation You Want To Perform :"))
repeter = 1
while(repeter):
    operation_No = int(input("Enter The Operation No You Want To Perform :"))
    print()

    match operation_No:
        case 1:
            CreateDatabase()
            print("DataBase Created Sucessfully------->")
        case 2 :
            Add_Employee(EmployeeNo)
            EmployeeNo+=1
            print("Employee Added Sucessfully--------->")
        case 3 :
            PromoteEmployee()
            print("Employee Promoted Sucessfully--------->")
        case 4:
            Remove_Employee()
            print("Employee Removed Sucessfully--------->")
        case 5:
            DisplayEmployeeData()
        case 6:
            DisplayWholeDataSet()
        case 7: 
            DeleteDataBase()
            repeter = 0
        case 8 :
            repeter = 0
            workbook.save("EmployeeDataBase.xlsx")
    print()
    print()



